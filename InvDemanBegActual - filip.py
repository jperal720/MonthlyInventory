from openpyxl import load_workbook
from datetime import date 
import datetime
import shutil as sh
import os

##THE FORMAT FOR THESE 2 FILES MUST BE THE SAME!!!
def getCurrentMonth():
    today = datetime.date.today()
    first = today.replace(day=1)
    currentMonth = first.strftime("%Y_%m")

    return currentMonth

def getLastMonth():
    today = datetime.date.today()
    first = today.replace(day=1)
    lastMonth = (first - datetime.timedelta(days=1)).strftime("%Y_%m")
    
    return lastMonth

inventoryTemplateFileDest = r"C:\Users\jokal\Documents\GitHub\MonthlyInventory\InventoryTemplate.xlsx"
oldExcelFileDest = r"C:\Users\jokal\Documents\GitHub\MonthlyInventory\Inventory" + str(getLastMonth()) + ".xlsx"
newExcelFileDest = r"C:\Users\jokal\Documents\GitHub\MonthlyInventory\Inventory" + str(getCurrentMonth()) + ".xlsx"

def moveValues(template, oldFile, newFile):
    """This function copies the needed values from the old file
    to the new monthly excel file"""

    oldWorkBook = load_workbook(filename= oldFile, data_only=True)
    oldWorkBook.sheetnames

    newWorkBook = load_workbook(filename= template)
    newWorkBook.sheetnames

    oldInventory = oldWorkBook['Inventory']
    newInventory = newWorkBook['Inventory']

    #From Sold Qty to Demand
    i = 2 
    while oldInventory.cell(row=i, column=1).value != None:
        newInventory.cell(row= i, column= 10).value = oldInventory.cell(row= i, column= 4).value
        i += 1
    
    #From Stock to Inv Beg
    i = 2
    while oldInventory.cell(row=i, column=1).value != None:
        newInventory.cell(row= i, column= 6).value = oldInventory.cell(row= i, column= 5).value
        i += 1

    newWorkBook.save(newFile)

# print(oldExcelFileDest)
moveValues(inventoryTemplateFileDest, oldExcelFileDest, newExcelFileDest)



